# #This takes in an excel-file and searchers for duplicates, to see the moves between 2021-01-01 and 2021-0-3-31 and makes sure
# #to weight this as per the move so if it's moved on day 25% and 75% days are remaning it will count 25% of the picks at the first location
# #and vise versa
# #Will need order history, 
import openpyxl
from openpyxl import load_workbook
from numpy.core.defchararray import isalpha, replace, strip
from datetime import datetime
from _datetime import date
#filepath for orderhistory
filepath = '/Users/eddijohansson/Desktop/Scania/MS order history .xlsx'
workbook_h = load_workbook(filepath)
sheet = workbook_h['SAMLING']
#Laddar in part number och hur många som beställts i en dict med tuples där part_no är key.
part_no_amount = {}
def norm(list):
    c = datetime.strptime('310321','%d%m%y') - datetime.strptime('010121','%d%m%y') 
    tal = []
    procent = 0
    negativ = [] 
    i = 0
    list.sort()
    for a in list:
        if int(a)>0:
            tal.append(a)
            procent = 1
        else:
            negativ.append(a)
    if procent == 1:
        tal.append(0)
        tal.append(89)
        tal.sort()
        last = 0
        for a in tal:
            try:
                tal[i] = a - last
                last += tal[i]
                i+=1
            except:
                i +=1
        i=0        
        while i < len(tal):
            tal[i] = float((tal[i])/89)
            i += 1
        tal.pop(0)
    else:
        tal = 1
    return tal

for col in sheet['A']:
    if not isalpha(str(col.value)):
        if strip(str(col.value)) != "":
            if not int(col.value) in part_no_amount:
                part_no_amount[col.value] = col.offset(0,1).value
            else:
                part_no_amount[col.value] += col.offset(0,1).value
#filepath for INVENTORY SNAPSHOT
filepath1 = '/Users/eddijohansson/Desktop/Scania/Inventory snapshot - MS.xlsx'
workbook_i = load_workbook(filepath1)
sheet_i = workbook_i['SAMLING']
part_no_days = {}
part_no_location = {}
#Laddar in part number och hur många som beställts i en dict med tuples där part_no är key. date är i 'E' och part_no är i 
a = 0

dat = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06','JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'}
for e in sheet_i['B']:
    if int(e.offset(0,1).value) not in part_no_days:
        string = str(e.offset(0,3).value)
        string = string.replace((str(e.offset(0,3).value)[3:6]), str(dat[(str(e.offset(0,3).value)[3:6])]))
        part_no_days[e.offset(0,1).value] = [string]
        part_no_location[e.offset(0,1).value] = [str(e.value)]
    else:
        string = str(e.offset(0,3).value)
        string = string.replace((str(e.offset(0,3).value)[3:6]), str(dat[(str(e.offset(0,3).value)[3:6])]))
        part_no_days[e.offset(0,1).value].append(string) 
        part_no_location[e.offset(0,1).value].append(str(e.value)) 
count = 0
for e in part_no_days:
    if len(part_no_days[e]) > 1: 
        i = 0
        while i < len(part_no_days[e]):
            string = part_no_days[e][i].replace("-","")
            string = datetime.strptime(string,'%d%m%y')
            c = datetime.strptime('010121','%d%m%y')
            b = datetime.strptime('310321','%d%m%y')
            if (string<=b):
                count += 1
                diff = string - c
                #Gives us days prior to the period as close to 0 is the pallets place at the begining
                #A positive value inicates how many days into the period it was moved e.g. 83 first move 84 second move diff = 1.
                part_no_days[e][i] = (diff.days)
            else:
                #Indicates that the data is not of interest
                part_no_days[e][i] = 'after'
            i += 1
    else:
        #Indicates that the pallet hasen't been moved at all - thus its place
        part_no_days[e][0] = 'place'

##Will have to sort now, and that means that everything that is classified as 'after' isn't of interest. Everything as 'place' is its 
#only place. Then the dynamics will come into play.
location_picks = {}
for e in part_no_days:
    i = 0
    while i < len(list(part_no_days[e])):
        print(list(part_no_days[e]))
        if part_no_days[e][i] == 'after' and len(part_no_days[e]) == 1:
            print('1')
            try:
                location_picks[part_no_location[e][i]] += part_no_amount[e]
                i+=1 
            except:
                i+=1 
        elif len(list(part_no_days[e])) == 1:
            print('2')
            try: 
                location_picks[part_no_location[e][i]] += part_no_amount[e]
                i+=1   
            except:
                i+=1    
        elif a == 0 and len(list(part_no_days[e])) > 1:
            print('3')
            part_no_location[e].pop(i)
            part_no_days[e].pop(i)
            i=0
        elif 'after' not in part_no_days[e] and len(part_no_days[e])>1:
            print('4')
            part_no_days[e], part_no_location[e] = zip(*sorted(zip(part_no_days[e], part_no_location[e])))
            listan = norm(list(part_no_days[e]))
            try:
                location_picks[part_no_location[e][i]] += part_no_amount[e]*listan[i]
                i+=1
            except:
                i+= 1   
        elif 'after' in part_no_days[e]:
            try:
                location_picks[part_no_location[e][i]] += part_no_amount[e]
            except:
                i += 1
        else:
            print('wtf')
            i+=1   

filepath3 = '/Users/eddijohansson/Desktop/Scania/OUTPUT/ungabunga.xlsx'
workbook_l = load_workbook(filepath3)
sheet_l = workbook_l['S']

r = 2
for a in location_picks:
    sheet_l.cell(row=r, column=1).value = a
    sheet_l.cell(row=r, column=2).value = location_picks[a]
    r += 1
workbook_l.save(filepath3)



