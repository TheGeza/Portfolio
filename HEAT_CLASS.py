# Created at LTH (Lunds Tekniska Högskola) for the course Warehousing - by Eddi Leino Johansson
#The script is takes two generic Excel-sheet and saves certain data points such as part number, # orders, placement, 
# and date of move. 
# Then generates a new file which can be converted to a heat map, if the user has the lay-out. 

class art:
    def __init__(self, nummer, location, days_spent, plock_tot):
        self.nummer = nummer
        self.location = location
        self.days_spent = days_spent
        self.plock_tot = plock_tot
        self.plock_at_place = [plock_tot]    
    def printa(self):
        print('')
        print(self.nummer)
        print(self.location)
        print(self.plock_tot)
        print(self.days_spent)
        print('')
    def plock_totc(self):
        return int(self.plock_tot)
    def num(self):
        return nummer
    def adda(self, plock_at_place):
        self.plock_at_place.append(plock_at_place)
        self.plock_tot += plock_at_place
    def sort(self):
         #part_no_days[e], part_no_location[e] = zip(*sorted(zip(part_no_days[e], part_no_location[e])))
        if all(element == 'after' for element in self.days_spent):
            i = 1
            while i < len(self.days_spent):
                self.location.pop(i)
                self.days_spent.pop(i)
        elif 'after' in self.days_spent and len(self.days_spent)>1:
            i = 0
            while i < len(self.days_spent):
                if(self.days_spent[i] == 'after'):
                    self.location.pop(i)
                    self.days_spent.pop(i)
                    i = 0
                i+= 1
        elif 'place' not in self.days_spent:
            self.days_spent, self.location = zip(*sorted(zip(self.days_spent, self.location)))
    def norm(self):
        c = datetime.strptime('310321','%d%m%y') - datetime.strptime('010121','%d%m%y') 
        tal = []
        procent = 0
        negativ = [] 
        i = 0
        for a in self.days_spent:
            if isalpha(str(a)):
                del self.days_spent
                self.days_spent = [1]
            elif int(a)>0:
                tal.append(a)
                procent = 1
        if procent == 1 and len(self.days_spent) == 1:
            del self.days_spent
            self.days_spent = [1]
        elif procent == 1 and len(self.days_spent) > 1:
            tal.append(89)
            tal.sort()
            last = 0
            for a in tal:
                try:
                    tal[i] = a - last
                    last += tal[i]
                    i+=1
                except:
                    i +=1
            i=0        
            while i < len(tal):
                tal[i] = float((tal[i])/89)
                i += 1
            del self.days_spent
            self.days_spent = tal
        else:
            del self.days_spent
            self.days_spent = [1]
    def plock_procent_ammount(self):
        if len(self.days_spent) > 1:
            asd = [x * self.plock_tot for x in self.days_spent]
            return asd
        else:
            asd = [self.plock_tot]
        return asd
    def plats(self):
        return list(self.location)
    def pplock (self):
        return list(self.plock_tot)


import openpyxl
from openpyxl import load_workbook
from numpy.core.defchararray import isalpha, replace, strip
from datetime import datetime
from _datetime import date


def load_hyllor_MS():
    #filepath for orderhistory
    filepath = '/Users/eddijohansson/Desktop/Scania/MS order history .xlsx'
    workbook_h = load_workbook(filepath)
    sheet = workbook_h['SAMLING']
    #Loades in part number and # of orders in a dict with tuples where part_no is the key.
    part_no_amount = {}
    for col in sheet['A']:
        if not isalpha(str(col.value)):
            if strip(str(col.value)) != "":
                if not int(col.value) in part_no_amount:
                    part_no_amount[col.value] = col.offset(0,1).value
                else:
                    part_no_amount[col.value] += col.offset(0,1).value
    #filepath for INVENTORY SNAPSHOT
    filepath1 = '/Users/eddijohansson/Desktop/Scania/Inventory snapshot - MS.xlsx'
    workbook_i = load_workbook(filepath1)
    sheet_i = workbook_i['SAMLING']
    part_no_days = {}
    part_no_location = {}
    a = 0

    dat = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06','JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'}
    for e in sheet_i['B']:
        if int(e.offset(0,1).value) not in part_no_days:
            string = str(e.offset(0,3).value)
            print(string)
            string = string.replace((str(e.offset(0,3).value)[3:6]), str(dat[(str(e.offset(0,3).value)[3:6])]))
            print(string)
            part_no_days[e.offset(0,1).value] = [string]
            part_no_location[e.offset(0,1).value] = [str(e.value)]
        else:
            string = str(e.offset(0,3).value)
            string = string.replace((str(e.offset(0,3).value)[3:6]), str(dat[(str(e.offset(0,3).value)[3:6])]))
            part_no_days[e.offset(0,1).value].append(string) 
            part_no_location[e.offset(0,1).value].append(str(e.value)) 
    count = 0
    for e in part_no_days:
        if len(part_no_days[e]) > 1: 
            i = 0
            while i < len(part_no_days[e]):
                string = part_no_days[e][i].replace("-","")
                string = datetime.strptime(string,'%d%m%y')
                c = datetime.strptime('010121','%d%m%y')
                b = datetime.strptime('310321','%d%m%y')
                if (string<=b):
                    count += 1
                    diff = string - c
                    #Gives us days prior to the period as close to 0 is the pallets place at the begining
                    #A positive value inicates how many days into the period it was moved e.g. 83 first move 84 second move diff = 1.
                    part_no_days[e][i] = (diff.days)
                else:
                    #Indicates that the data is not of interest
                    part_no_days[e][i] = 'after'
                i += 1
        else:
            #Indicates that the pallet hasen't been moved at all - thus its place
            part_no_days[e][0] = 'place'

    ##Will have to sort now, and that means that everything that is classified as 'after' isn't of interest. Everything as 'place' is its 
    #only place. Then the dynamics will come into play.

    lista_art = []
    for a in part_no_amount:
        if (any(x.num == a for x in lista_art)):
            a.adda(part_no_amount[a])
        else:
            try:
                lista_art.append(art(str(a),part_no_location[a], part_no_days[a], part_no_amount[a]))
            except:
                lista_art.append(art(str(a), ['unkown'], ['unkown'], part_no_amount[a]))

    tot = 0
    for a in lista_art:
        a.sort()
        a.norm()
        # if sum(a.plock_procent_ammount()) > 100:
        #     print(a.plats())
        #     print(a.pplock())
        #     print(a.plock_procent_ammount())
        # tot += sum(a.plock_procent_ammount())




    filepath3 = '/Users/eddijohansson/Desktop/Scania/OUTPUT/ungabunga.xlsx'
    workbook_l = load_workbook(filepath3)
    sheet_l = workbook_l['S']

    r = 0
    plats = []
    count = [] 
    plats_ex = []
    count_ex = []
    for a in lista_art:
        plats = (list(a.plats()))
        count = (list(a.plock_procent_ammount()))
        i = len(count) - 1
        j = len(plats) - 1 
        for a in plats:
            if list(plats)[j] not in plats_ex and i>-1:
                plats_ex.append(list(plats).pop(j))
                count_ex.append(list(count).pop(i))
                i = i - 1
                j = j - 1 
            elif i>-1:
                r = plats_ex.index(list(plats).pop(j))
                count_ex[r] += list(count).pop(i)
                i = i - 1
                j = j - 1 

    r= 2
    for a in plats_ex:
        sheet_l.cell(row=r, column=1).value = a
        r += 1
    r = 2
    for b in count_ex:
        sheet_l.cell(row=r, column=6).value = b
        r += 1
    workbook_l.save(filepath3)
